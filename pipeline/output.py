"""
pipeline/output.py — Stage 8: Output Generation

Produces three outputs from every pipeline run:

1. JSON file  — timestamped, API-spec compliant, saved to outputs/
2. Excel file — formatted workbook with 4 sheets:
                Portfolio | All Scores | Trade Signals | Run Log
3. HTML dashboard — self-contained interactive page

Also saves the portfolio state file (latest_portfolio.json)
used by signals.py for next month's HOLD/EXIT detection.
"""

import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd

sys.path.insert(0, str(Path(__file__).parent.parent))
import config

log = logging.getLogger(__name__)


# ── Helpers ────────────────────────────────────────────────────────────────

def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def _run_label() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _market_regime(scored_df: pd.DataFrame) -> str:
    """Estimate market regime from scored universe."""
    if scored_df.empty:
        return "unknown"
    above_ma = scored_df["above_sma200"].mean() if "above_sma200" in scored_df.columns else 0.5
    avg_mom  = scored_df["ret_3m"].mean() if "ret_3m" in scored_df.columns else 0
    if above_ma >= 0.65 and avg_mom > 0.03:
        return "bullish"
    elif above_ma <= 0.40 or avg_mom < -0.05:
        return "bearish"
    return "neutral"


def _risk_summary(portfolio: list, signal_result: dict) -> dict:
    """Compute portfolio-level risk summary."""
    vols = [p["risk_proxy_vol"] for p in portfolio if p.get("risk_proxy_vol") is not None]
    avg_vol = sum(vols) / len(vols) if vols else None

    # Simplified max drawdown estimate: 2× avg volatility × sqrt(1/12) for 1 month
    import math
    mdd_est = f"~{avg_vol * 2 * math.sqrt(1/12) * 100:.1f}%" if avg_vol else "N/A"
    vol_level = "High" if avg_vol and avg_vol > 0.35 else "Moderate" if avg_vol and avg_vol > 0.20 else "Low"

    exit_count = signal_result.get("signal_summary", {}).get("exit", 0)
    buy_count  = signal_result.get("signal_summary", {}).get("buy", 0)

    return {
        "max_drawdown_estimate": mdd_est,
        "volatility_level":     vol_level,
        "avg_position_vol":     f"{avg_vol*100:.1f}%" if avg_vol else "N/A",
        "notes": (
            f"{exit_count} position(s) exited, {buy_count} new position(s) added this month. "
            "Equal-weight allocation, monthly rebalance. "
            "Past factor performance does not guarantee future returns."
        ),
    }


# ── Output 1: JSON ─────────────────────────────────────────────────────────

def save_json(
    all_stage_results: dict,
    run_label: str,
) -> Path:
    """Save the final API-spec JSON file."""
    sel_result  = all_stage_results["selection"]
    port_result = all_stage_results["portfolio"]
    sig_result  = all_stage_results["signals"]
    feat_result = all_stage_results["features"]
    ing_result  = all_stage_results["ingestion"]

    scored_df     = all_stage_results.get("scoring", {}).get("scored", pd.DataFrame())
    regime_result = all_stage_results.get("regime", {}) or {}
    # Use live regime from regime.py if available, otherwise infer from scored data
    if regime_result.get("regime"):
        regime_label = regime_result["regime"]
        # Map to output.py's convention (bull -> bullish etc.)
        regime = {"bull": "bullish", "neutral": "neutral", "bear": "bearish"}.get(
            regime_label, regime_label
        )
    else:
        regime = _market_regime(scored_df)

    output = {
        "timestamp":       _timestamp(),
        "run_label":       run_label,
        "market_regime":   regime,
        "regime_detail": {
            "regime":           regime_result.get("regime", "unknown"),
            "vix_current":      regime_result.get("vix_current"),
            "spx_price":        regime_result.get("spx_price"),
            "spx_200ma":        regime_result.get("spx_200ma"),
            "spx_vs_200ma_pct": regime_result.get("spx_vs_200ma_pct"),
            "active_top_n":     regime_result.get("active_top_n"),
            "notes":            regime_result.get("notes", ""),
        },
        "data_sources":    ing_result.get("data_sources", []),
        "date_range":      ing_result.get("date_range", {}),
        "universe_size":   feat_result.get("ticker_count", 0),
        "top_10_stocks":   sel_result.get("top_10_stocks", []),
        "portfolio":       port_result.get("portfolio", []),
        "trade_signals":   sig_result.get("all_signals", []),
        "risk_summary":    _risk_summary(port_result.get("portfolio", []), sig_result),
        "signal_summary":  sig_result.get("signal_summary", {}),
        "filter_funnel":   all_stage_results.get("filters", {}).get("filter_funnel", []),
    }

    path = config.OUTPUT_DIR / f"portfolio_{run_label}.json"
    with open(path, "w") as f:
        json.dump(output, f, indent=2, default=str)
    log.info(f"  JSON saved → {path.name}")
    return path


def save_portfolio_state(portfolio: list) -> None:
    """Overwrite the latest portfolio state for next month's HOLD/EXIT detection."""
    with open(config.PORTFOLIO_STATE_FILE, "w") as f:
        json.dump({"portfolio": portfolio}, f, indent=2)
    log.info(f"  Portfolio state saved → {config.PORTFOLIO_STATE_FILE.name}")


# ── Output 2: Excel ────────────────────────────────────────────────────────

def save_excel(all_stage_results: dict, run_label: str) -> Path:
    """Save a formatted 4-sheet Excel workbook."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Alignment, Border, Font, PatternFill,
                                     Side)
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl not installed — skipping Excel output")
        return None

    port_result = all_stage_results["portfolio"]
    sel_result  = all_stage_results["selection"]
    sig_result  = all_stage_results["signals"]
    feat_result = all_stage_results["features"]

    wb = Workbook()

    # ── Styles ──────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="003366")
    HDR_FONT  = Font(color="FFFFFF", bold=True)
    ALT_FILL  = PatternFill("solid", fgColor="EBF0F7")
    BUY_FILL  = PatternFill("solid", fgColor="D6F5D6")
    HOLD_FILL = PatternFill("solid", fgColor="FFF9DB")
    EXIT_FILL = PatternFill("solid", fgColor="FFE0E0")
    THIN      = Side(style="thin", color="CCCCCC")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def style_header(ws, row_num, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # ── Sheet 1: Portfolio ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Portfolio"
    headers1 = ["Rank","Ticker","Name","Sector","Weight","Score","Price","Entry Price","Entry Date","Exp Return","Vol","Trend","Momentum"]
    ws1.append(headers1)
    style_header(ws1, 1, len(headers1))

    for i, p in enumerate(port_result.get("portfolio", []), 2):
        # Pull entry_price/entry_date from signals if available, else from portfolio
        sig_map = {s["ticker"]: s for s in all_stage_results.get("signals", {}).get("trade_signals", [])}
        sig_item = sig_map.get(p["ticker"], {})
        entry_price = sig_item.get("entry_price") or p.get("entry_price") or p["current_price"]
        entry_date  = sig_item.get("entry_date", "")
        row = [
            p["rank"], p["ticker"], p["name"], p.get("sector",""),
            p["weight_pct"], p["score"], p["current_price"],
            entry_price, entry_date,
            f"{(p['expected_return_proxy'] or 0)*100:.1f}%" if p.get("expected_return_proxy") else "N/A",
            f"{(p['risk_proxy_vol'] or 0)*100:.1f}%" if p.get("risk_proxy_vol") else "N/A",
            p["signals"]["trend"], p["signals"]["momentum"],
        ]
        ws1.append(row)
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for c in range(1, len(headers1)+1):
            ws1.cell(row=i, column=c).border = BORDER
            if i % 2 == 0:
                ws1.cell(row=i, column=c).fill = ALT_FILL
    auto_width(ws1)

    # ── Sheet 2: All Scores ──────────────────────────────────────────────
    ws2 = wb.create_sheet("All Scores")
    scored_df = all_stage_results.get("scoring", {}).get("scored", pd.DataFrame())
    if not scored_df.empty:
        cols2 = ["ticker","name","composite_score","score_momentum","score_trend","score_quality","score_volatility",
                 "ret_3m","ret_6m","ret_12m","vol_60d","above_sma200","rsi_14"]
        sub = scored_df[[c for c in cols2 if c in scored_df.columns]].sort_values("composite_score", ascending=False)
        ws2.append(list(sub.columns))
        style_header(ws2, 1, len(sub.columns))
        for i, (_, row) in enumerate(sub.iterrows(), 2):
            ws2.append([round(v, 4) if isinstance(v, float) else v for v in row.tolist()])
            for c in range(1, len(sub.columns)+1):
                ws2.cell(row=i, column=c).border = BORDER
                if i % 2 == 0:
                    ws2.cell(row=i, column=c).fill = ALT_FILL
    auto_width(ws2)

    # ── Sheet 3: Trade Signals ───────────────────────────────────────────
    ws3 = wb.create_sheet("Trade Signals")
    headers3 = ["Action","Ticker","Name","Weight","Score","Trend","Momentum","Entry Rationale","Risk Note"]
    ws3.append(headers3)
    style_header(ws3, 1, len(headers3))
    for i, s in enumerate(sig_result.get("all_signals", []), 2):
        row3 = [
            s["action"], s["ticker"], s.get("name",""),
            f"{s['weight']*100:.0f}%" if s.get("weight") else "0%",
            s.get("composite_score",""),
            s.get("signals",{}).get("trend",""), s.get("signals",{}).get("momentum",""),
            s.get("entry_rationale",""), s.get("risk_note",""),
        ]
        ws3.append(row3)
        action_fill = BUY_FILL if s["action"]=="BUY" else HOLD_FILL if s["action"]=="HOLD" else EXIT_FILL
        for c in range(1, len(headers3)+1):
            cell = ws3.cell(row=i, column=c)
            cell.fill = action_fill
            cell.border = BORDER
    ws3.column_dimensions["H"].width = 50
    ws3.column_dimensions["I"].width = 45
    auto_width(ws3)

    # ── Sheet 4: Run Log ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Run Log")
    ws4.append(["Run Summary"])
    ws4["A1"].font = Font(bold=True, size=14)
    ws4.append([])
    funnel = all_stage_results.get("filters", {}).get("filter_funnel", [])
    ws4.append(["Stage","Count","Removed"])
    style_header(ws4, 3, 3)
    for step in funnel:
        ws4.append([step["step"], step["count"], step["removed"]])
    ws4.append([])
    ws4.append(["Run label", run_label])
    ws4.append(["Timestamp", _timestamp()])
    ws4.append(["Universe size", feat_result.get("ticker_count",0)])
    ws4.append(["Selected", sel_result.get("ticker_count",0)])
    auto_width(ws4)

    path = config.OUTPUT_DIR / f"portfolio_{run_label}.xlsx"
    wb.save(path)
    log.info(f"  Excel saved → {path.name}")
    return path



# ── Paper trading + stop-loss data loaders ────────────────────────────────

def _load_paper_trading_data() -> dict:
    """Load latest snapshot from paper_trading_log.json. Returns {} if missing."""
    try:
        log_path = Path(getattr(config, "DATA_DIR", "data")) / "paper_trading_log.json"
        if not log_path.exists():
            return {}
        data = json.loads(log_path.read_text(encoding="utf-8"))
        if not data:
            return {}
        return data[-1]          # most recent snapshot
    except Exception:
        return {}


def _load_stop_loss_data() -> list:
    """Load stop_loss_log.json. Returns list of position dicts."""
    try:
        sl_path = config.OUTPUT_DIR / "stop_loss_log.json"
        if not sl_path.exists():
            return []
        raw = json.loads(sl_path.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            return raw
        return []
    except Exception:
        return []


# ── Output 3: HTML Dashboard ──────────────────────────────────────────────

def save_dashboard(all_stage_results: dict, run_label: str) -> Path:
    """Save a self-contained interactive HTML dashboard."""
    port_result = all_stage_results["portfolio"]
    sig_result  = all_stage_results["signals"]
    ing_result  = all_stage_results["ingestion"]
    scored_df     = all_stage_results.get("scoring", {}).get("scored", pd.DataFrame())
    regime_result = all_stage_results.get("regime", {}) or {}
    if regime_result.get("regime"):
        regime = {"bull": "bullish", "neutral": "neutral", "bear": "bearish"}.get(
            regime_result["regime"], regime_result["regime"]
        )
    else:
        regime = _market_regime(scored_df)
    risk        = _risk_summary(port_result.get("portfolio", []), sig_result)
    portfolio   = port_result.get("portfolio", [])
    signals     = sig_result.get("all_signals", [])

    funnel      = all_stage_results.get("filters", {}).get("filter_funnel", [])

    # Paper trading + stop-loss data
    pt  = _load_paper_trading_data()
    sls = {e["ticker"]: e for e in _load_stop_loss_data()}

    regime_color = {"bullish": "#2ecc71", "bearish": "#e74c3c", "neutral": "#f39c12"}.get(regime, "#95a5a6")

    # Build regime detail string for the banner
    vix_str = f"VIX {regime_result['vix_current']:.1f}" if regime_result.get("vix_current") else ""
    spx_str = (f"SPX {regime_result['spx_vs_200ma_pct']:+.1f}% vs 200MA"
               if regime_result.get("spx_vs_200ma_pct") is not None else "")
    regime_detail_str = "  |  ".join(s for s in [vix_str, spx_str] if s)

    # Portfolio rows
    port_rows = ""
    for p in portfolio:
        trend_badge = f'<span class="badge badge-{"bull" if p["signals"]["trend"]=="bullish" else "bear" if p["signals"]["trend"]=="bearish" else "neu"}">{p["signals"]["trend"]}</span>'
        mom_badge   = f'<span class="badge badge-{"strong" if p["signals"]["momentum"]=="strong" else "mod"}">{p["signals"]["momentum"]}</span>'
        exp_ret     = f'{p["expected_return_proxy"]*100:.1f}%' if p.get("expected_return_proxy") else "N/A"
        vol         = f'{p["risk_proxy_vol"]*100:.1f}%' if p.get("risk_proxy_vol") else "N/A"
        port_rows += f"""
        <tr>
          <td><strong>{p['rank']}</strong></td>
          <td><strong>{p['ticker']}</strong></td>
          <td>{p['name']}</td>
          <td>{p.get('sector','')}</td>
          <td>{p['weight_pct']}</td>
          <td>{p['score']:.4f}</td>
          <td>${p['current_price']:.2f}</td>
          <td>{exp_ret}</td>
          <td>{vol}</td>
          <td>{trend_badge}</td>
          <td>{mom_badge}</td>
        </tr>"""

    # Signal rows
    sig_rows = ""
    for s in signals:
        action_class = {"BUY": "action-buy", "HOLD": "action-hold", "EXIT": "action-exit"}.get(s["action"], "")
        sig_rows += f"""
        <tr>
          <td><span class="action {action_class}">{s['action']}</span></td>
          <td><strong>{s['ticker']}</strong></td>
          <td>{s.get('name','')}</td>
          <td>{s.get('entry_rationale','')[:80]}</td>
          <td>{s.get('risk_note','')[:60]}</td>
        </tr>"""

    # Funnel rows
    funnel_rows = ""
    for step in funnel:
        funnel_rows += f"<tr><td>{step['step']}</td><td>{step['count']}</td><td style='color:#e74c3c'>-{step['removed']}</td></tr>"

    # Score bars for portfolio
    score_bars = ""
    for p in portfolio:
        score_bars += f"""
        <div class="score-bar-row">
          <div class="score-label">{p['ticker']}</div>
          <div class="score-bar-bg">
            <div class="score-bar-fill" style="width:{p['score']*100:.0f}%"></div>
          </div>
          <div class="score-val">{p['score']:.3f}</div>
        </div>"""


    # ── Paper Trading Progress section ────────────────────────────────────
    pt_day    = pt.get("day_number", 1)
    pt_total  = int(getattr(config, "PAPER_TRADING_MONTHS", 3)) * 30
    pt_pct    = min(100, int(pt_day / pt_total * 100)) if pt_total else 0
    pt_val    = pt.get("total_portfolio_value", 0) or 0
    pt_ret    = pt.get("total_return_pct", 0) or 0
    pt_sharpe = pt.get("sharpe_ratio") or 0
    pt_mdd    = pt.get("max_drawdown_pct", 0) or 0
    pt_alpha  = pt.get("alpha_pct") or 0
    pt_bench  = getattr(config, "BENCHMARK_TICKER", "SPY")
    pt_start  = getattr(config, "PAPER_TRADING_START_DATE", "")
    pt_regime = (pt.get("regime") or "unknown").upper()
    min_obs   = getattr(config, "MIN_FEEDBACK_OBSERVATIONS", 25)
    accum_obs = 0
    try:
        from pipeline import feedback as _fb
        accum_obs = _fb.count_accumulated_observations()
    except Exception:
        pass
    accum_pct = min(100, int(accum_obs / max(min_obs, 1) * 100))

    ret_color = "#2ecc71" if pt_ret >= 0 else "#e74c3c"
    alp_color = "#2ecc71" if pt_alpha >= 0 else "#e74c3c"
    mdd_color = "#e74c3c" if pt_mdd < -5 else ("#f39c12" if pt_mdd < -2 else "#7a8aaa")
    shp_color = "#2ecc71" if pt_sharpe >= 1 else ("#f39c12" if pt_sharpe >= 0 else "#e74c3c")
    wt_status = ("Weight updates ACTIVE" if accum_obs >= min_obs
                 else str(min_obs - accum_obs) + " more observations needed")
    wt_icon   = "checkmark" if accum_obs >= min_obs else "hourglass"

    # Build position rows
    pos_rows_html = ""
    for pos in pt.get("positions", []):
        tk      = pos.get("ticker", "")
        ent     = pos.get("entry_price", 0) or 0
        cur     = pos.get("current_price", 0) or 0
        pnl     = pos.get("pnl_pct", 0) or 0
        pnl_eur = pos.get("pnl_eur", 0) or 0
        sl_e    = sls.get(tk, {})
        stop    = sl_e.get("stop_price") or sl_e.get("stop") or 0
        if cur and stop:
            sl_gap  = (cur - stop) / cur * 100
            bar_pct = min(100, max(0, int(sl_gap * 4)))
            if sl_gap < 5:
                bar_col = "#e74c3c"
                txt_col = "#e74c3c"
            elif sl_gap < 10:
                bar_col = "#f39c12"
                txt_col = "#f39c12"
            else:
                bar_col = "#2ecc71"
                txt_col = "#2ecc71"
            sl_cell = (
                '<span style="font-size:11px;color:' + txt_col + '">'
                + '{:.1f}% above stop</span> '.format(sl_gap)
                + '<span class="sl-bar-outer"><span class="sl-bar-inner" style="width:'
                + str(bar_pct) + '%;background:' + bar_col + '"></span></span>'
            )
        else:
            sl_cell = '<span style="color:#7a8aaa">N/A</span>'
        pnl_cls = "pos-gain" if pnl > 0 else ("pos-loss" if pnl < 0 else "pos-flat")
        arrow   = "&#9650;" if pnl > 0 else ("&#9660;" if pnl < 0 else "&mdash;")
        pos_rows_html += (
            "<tr>"
            + "<td><strong>" + tk + "</strong></td>"
            + "<td>${:.2f}</td>".format(ent)
            + "<td>${:.2f}</td>".format(cur)
            + '<td class="' + pnl_cls + '">' + arrow + " {:.2f}%</td>".format(pnl)
            + '<td class="' + pnl_cls + '">&euro;{:+.2f}</td>'.format(pnl_eur)
            + "<td>" + sl_cell + "</td>"
            + "</tr>"
        )

    if not pos_rows_html:
        pos_rows_html = ('<tr><td colspan="6" style="color:#7a8aaa;text-align:center">'
                         'No positions tracked yet &mdash; run main.py to initialise</td></tr>')

    if pt:
        progress_html = (
            '\n  <!-- Paper Trading Progress -->\n'
            '  <div class="section">\n'
            '    <h2>&#128200; Paper Trading Progress &mdash; Day '
            + str(pt_day) + ' of ' + str(pt_total) + '</h2>\n'
            '    <div class="prog-label">' + str(pt_day) + ' / ' + str(pt_total)
            + ' days (' + str(pt_pct) + '% complete)'
            + ' &nbsp;|&nbsp; Started ' + str(pt_start)
            + ' &nbsp;|&nbsp; Regime: ' + pt_regime + '</div>\n'
            '    <div class="prog-bar-outer"><div class="prog-bar-inner" style="width:'
            + str(pt_pct) + '%"></div></div>\n\n'
            '    <div class="kpi-row">\n'
            '      <div class="kpi-box"><div class="klabel">Portfolio Value</div>'
            '<div class="kvalue">&euro;{:.2f}</div>'.format(pt_val)
            + '<div class="ksub">starting &euro;1,000</div></div>\n'
            '      <div class="kpi-box"><div class="klabel">Total Return</div>'
            '<div class="kvalue" style="color:' + ret_color + '">{:+.2f}%</div>'.format(pt_ret)
            + '<div class="ksub">since paper start</div></div>\n'
            '      <div class="kpi-box"><div class="klabel">Alpha vs ' + pt_bench + '</div>'
            '<div class="kvalue" style="color:' + alp_color + '">{:+.2f}%</div>'.format(pt_alpha)
            + '<div class="ksub">excess return</div></div>\n'
            '      <div class="kpi-box"><div class="klabel">Sharpe / Max DD</div>'
            '<div class="kvalue" style="color:' + shp_color + '">{:.2f}</div>'.format(pt_sharpe)
            + '<div class="ksub" style="color:' + mdd_color + '">MDD {:.2f}%</div>'.format(pt_mdd)
            + '</div>\n    </div>\n\n'
            '    <table>\n'
            '      <thead><tr>'
            '<th>Ticker</th><th>Entry</th><th>Current</th>'
            '<th>P&amp;L %</th><th>P&amp;L &euro;</th><th>Stop-Loss Distance</th>'
            '</tr></thead>\n'
            '      <tbody>' + pos_rows_html + '</tbody>\n'
            '    </table>\n\n'
            '    <div style="margin-top:16px;font-size:12px;color:#7a8aaa">\n'
            '      <strong style="color:#a8c8f0">Factor Weight Learning:</strong>'
            ' &nbsp; ' + str(accum_obs) + ' / ' + str(min_obs) + ' observations'
            ' &nbsp; <span class="accum-bar-outer" style="width:120px">'
            '<span class="accum-bar-inner" style="width:' + str(accum_pct) + '%"></span></span>'
            ' &nbsp; ' + wt_status + '\n'
            '    </div>\n'
            '  </div>\n'
        )
    else:
        progress_html = (
            '\n  <div class="section">\n'
            '    <h2>&#128200; Paper Trading Progress</h2>\n'
            '    <p style="color:#7a8aaa;font-size:13px">No paper trading data yet.'
            ' Run <code>main.py</code> to start tracking.</p>\n'
            '  </div>\n'
        )


    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Investment Alpha Dashboard — {run_label}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #0f1117; color: #e0e6f0; }}
  .header {{ background: linear-gradient(135deg, #003366 0%, #0066cc 100%); padding: 28px 40px; }}
  .header h1 {{ font-size: 24px; font-weight: 700; letter-spacing: 1px; }}
  .header .meta {{ font-size: 13px; color: #a8c8f0; margin-top: 6px; }}
  .regime-badge {{ display: inline-block; padding: 4px 14px; border-radius: 20px; font-size: 13px; font-weight: 700;
                   background: {regime_color}; color: #fff; margin-left: 12px; vertical-align: middle; }}
  .container {{ padding: 28px 40px; max-width: 1400px; }}
  .cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 28px; }}
  .card {{ background: #1a2035; border-radius: 10px; padding: 20px; border-left: 4px solid #0066cc; }}
  .card .label {{ font-size: 11px; color: #7a8aaa; text-transform: uppercase; letter-spacing: 1px; }}
  .card .value {{ font-size: 26px; font-weight: 700; margin-top: 6px; color: #e0e6f0; }}
  .card .sub {{ font-size: 12px; color: #7a8aaa; margin-top: 4px; }}
  .section {{ background: #1a2035; border-radius: 10px; padding: 22px; margin-bottom: 22px; }}
  .section h2 {{ font-size: 15px; font-weight: 700; color: #a8c8f0; margin-bottom: 16px; text-transform: uppercase; letter-spacing: 1px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  th {{ background: #003366; color: #a8c8f0; padding: 9px 12px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #232b40; }}
  tr:hover td {{ background: #1e2845; }}
  .badge {{ padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }}
  .badge-bull {{ background: #1a3a2a; color: #2ecc71; }}
  .badge-bear {{ background: #3a1a1a; color: #e74c3c; }}
  .badge-neu  {{ background: #2a2a1a; color: #f39c12; }}
  .badge-strong {{ background: #1a3a2a; color: #2ecc71; }}
  .badge-mod  {{ background: #1a2a3a; color: #3498db; }}
  .action {{ padding: 3px 10px; border-radius: 4px; font-size: 11px; font-weight: 700; }}
  .action-buy  {{ background: #1a3a2a; color: #2ecc71; }}
  .action-hold {{ background: #2a2a1a; color: #f39c12; }}
  .action-exit {{ background: #3a1a1a; color: #e74c3c; }}
  .score-bar-row {{ display: flex; align-items: center; margin-bottom: 8px; }}
  .score-label {{ width: 60px; font-size: 13px; font-weight: 600; }}
  .score-bar-bg {{ flex: 1; background: #232b40; border-radius: 4px; height: 10px; margin: 0 12px; }}
  .score-bar-fill {{ background: linear-gradient(90deg, #0066cc, #00cc99); border-radius: 4px; height: 10px; transition: width 0.5s; }}
  .score-val {{ width: 45px; font-size: 12px; color: #7a8aaa; text-align: right; }}
  .two-col {{ display: grid; grid-template-columns: 1fr 1fr; gap: 22px; }}
  .risk-item {{ padding: 8px 0; border-bottom: 1px solid #232b40; font-size: 13px; }}
  .risk-label {{ color: #7a8aaa; font-size: 11px; }}
  .risk-value {{ font-weight: 600; }}

  .prog-bar-outer {{ background: #232b40; border-radius: 6px; height: 14px; margin: 10px 0 18px 0; }}
  .prog-bar-inner {{ background: linear-gradient(90deg, #0066cc, #00cc99); border-radius: 6px; height: 14px; }}
  .prog-label {{ font-size: 12px; color: #7a8aaa; margin-bottom: 4px; }}
  .kpi-row {{ display: grid; grid-template-columns: repeat(4,1fr); gap: 14px; margin-bottom: 20px; }}
  .kpi-box {{ background: #13192b; border-radius: 8px; padding: 14px 16px; border-top: 3px solid #0066cc; }}
  .kpi-box .klabel {{ font-size: 10px; color: #7a8aaa; text-transform: uppercase; letter-spacing: 1px; }}
  .kpi-box .kvalue {{ font-size: 22px; font-weight: 700; margin-top: 4px; }}
  .kpi-box .ksub {{ font-size: 11px; color: #7a8aaa; margin-top: 3px; }}
  .pos-gain {{ color: #2ecc71; }}
  .pos-loss {{ color: #e74c3c; }}
  .pos-flat {{ color: #7a8aaa; }}
  .sl-bar-outer {{ width: 90px; background: #232b40; border-radius: 4px; height: 8px; display:inline-block; vertical-align:middle; }}
  .sl-bar-inner {{ background: #2ecc71; border-radius: 4px; height: 8px; }}
  .sl-bar-warn  {{ background: #f39c12; }}
  .sl-bar-crit  {{ background: #e74c3c; }}
  .accum-bar-outer {{ background: #232b40; border-radius: 4px; height: 8px; display:inline-block; vertical-align:middle; }}
  .accum-bar-inner {{ background: #f39c12; border-radius: 4px; height: 8px; }}
  footer {{ text-align: center; color: #3a4a6a; font-size: 11px; padding: 20px; }}
</style>
</head>
<body>
<div class="header">
  <h1>📈 Investment Alpha Dashboard
    <span class="regime-badge">Market: {regime.upper()}</span>
  </h1>
  <div class="meta">Run: {run_label} &nbsp;|&nbsp; Universe: {ing_result.get('date_range',{}).get('start','')}&nbsp;→&nbsp;{ing_result.get('date_range',{}).get('end','')} &nbsp;|&nbsp; Data: {', '.join(ing_result.get('data_sources',[]))}
  {f'&nbsp;|&nbsp; {regime_detail_str}' if regime_detail_str else ''}</div>
</div>
<div class="container">

  <!-- KPI Cards -->
  <div class="cards">
    <div class="card">
      <div class="label">Stocks Selected</div>
      <div class="value">{len(portfolio)}</div>
      <div class="sub">from {all_stage_results.get('features',{}).get('ticker_count',0)} screened</div>
    </div>
    <div class="card">
      <div class="label">Portfolio Weight / Stock</div>
      <div class="value">{f"{portfolio[0]['weight']*100:.0f}%" if portfolio else "N/A"}</div>
      <div class="sub">equal weight</div>
    </div>
    <div class="card">
      <div class="label">Volatility Level</div>
      <div class="value">{risk['volatility_level']}</div>
      <div class="sub">avg vol {risk['avg_position_vol']}</div>
    </div>
    <div class="card">
      <div class="label">Max Drawdown Est.</div>
      <div class="value">{risk['max_drawdown_estimate']}</div>
      <div class="sub">1-month estimate</div>
    </div>
  </div>

  <!-- Portfolio Table -->
  <div class="section">
    <h2>🏆 Selected Portfolio</h2>
    <table>
      <thead><tr><th>#</th><th>Ticker</th><th>Name</th><th>Sector</th><th>Weight</th><th>Score</th><th>Price</th><th>Exp Return</th><th>Vol</th><th>Trend</th><th>Momentum</th></tr></thead>
      <tbody>{port_rows}</tbody>
    </table>
  </div>

  <div class="two-col">
    <!-- Score Bars -->
    <div class="section">
      <h2>📊 Composite Scores</h2>
      {score_bars}
    </div>

    <!-- Risk Summary -->
    <div class="section">
      <h2>⚠️ Risk Summary</h2>
      <div class="risk-item"><div class="risk-label">Max Drawdown Estimate</div><div class="risk-value">{risk['max_drawdown_estimate']}</div></div>
      <div class="risk-item"><div class="risk-label">Volatility Level</div><div class="risk-value">{risk['volatility_level']}</div></div>
      <div class="risk-item"><div class="risk-label">Avg Position Volatility</div><div class="risk-value">{risk['avg_position_vol']}</div></div>
      <div class="risk-item"><div class="risk-label">Notes</div><div class="risk-value" style="font-weight:400;font-size:12px">{risk['notes']}</div></div>
    </div>
  </div>

  <!-- Trade Signals -->
  <div class="section">
    <h2>📡 Trade Signals</h2>
    <table>
      <thead><tr><th>Action</th><th>Ticker</th><th>Name</th><th>Rationale</th><th>Risk Note</th></tr></thead>
      <tbody>{sig_rows}</tbody>
    </table>
  </div>

  <!-- Filter Funnel -->
  <div class="section">
    <h2>🔽 Filter Funnel</h2>
    <table>
      <thead><tr><th>Stage</th><th>Stocks</th><th>Removed</th></tr></thead>
      <tbody>{funnel_rows}</tbody>
    </table>
  </div>

{progress_html}
</div>
<footer>Investment Alpha Pipeline — Generated {_timestamp()} — For research purposes only. Not financial advice.</footer>
</body>
</html>"""

    path = config.OUTPUT_DIR / f"dashboard_{run_label}.html"
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info(f"  Dashboard saved → {path.name}")
    return path


# ── Stage 8 Orchestrator ──────────────────────────────────────────────────

def run(all_stage_results: dict) -> dict:
    """
    Stage 8: Generate all outputs.

    Returns:
        {
            "stage": "output",
            "status": "success" | "partial" | "failed",
            "json_path": Path,
            "excel_path": Path,
            "dashboard_path": Path,
            "run_label": str,
        }
    """
    log.info(f"\n{'='*50}")
    log.info("STAGE 8: Output Generation")
    log.info(f"{'='*50}")

    run_label = _run_label()
    outputs   = {"stage": "output", "run_label": run_label}

    # 1. Save portfolio state (for next run's BUY/HOLD/EXIT)
    portfolio = all_stage_results.get("portfolio", {}).get("portfolio", [])
    if portfolio:
        save_portfolio_state(portfolio)

    # 2. JSON
    try:
        outputs["json_path"] = save_json(all_stage_results, run_label)
    except Exception as e:
        log.error(f"  JSON output failed: {e}")
        outputs["json_path"] = None

    # 3. Excel
    try:
        outputs["excel_path"] = save_excel(all_stage_results, run_label)
    except Exception as e:
        log.error(f"  Excel output failed: {e}")
        outputs["excel_path"] = None

    # 4. HTML Dashboard
    try:
        outputs["dashboard_path"] = save_dashboard(all_stage_results, run_label)
    except Exception as e:
        log.error(f"  Dashboard output failed: {e}")
        outputs["dashboard_path"] = None

    success_count = sum(1 for k in ["json_path","excel_path","dashboard_path"] if outputs.get(k))
    outputs["status"] = "success" if success_count == 3 else "partial" if success_count > 0 else "failed"

    log.info(f"Stage 8 complete — {success_count}/3 outputs generated")
    return outputs


# ── Quick Test ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    from pipeline import ingestion, features, scoring, filters, selection, portfolio, signals

    print("\n=== Stage 8 Test: Output Generation ===")
    TEST_TICKERS = ["AAPL","MSFT","GOOGL","AMZN","NVDA","META","JPM","JNJ","V","UNH"]

    ing   = ingestion.run(tickers=TEST_TICKERS)
    feat  = features.run(ing)
    sc    = scoring.run(feat)
    filt  = filters.run(sc)
    sel   = selection.run(filt, top_n=5)
    port  = portfolio.run(sel)
    sigs  = signals.run(port, sel)

    all_results = {
        "ingestion": ing, "features": feat, "scoring": sc,
        "filters": filt, "selection": sel, "portfolio": port, "signals": sigs,
    }

    result = run(all_results)

    print(f"\nStatus         : {result['status']}")
    print(f"Run label      : {result['run_label']}")
    print(f"JSON path      : {result.get('json_path')}")
    print(f"Excel path     : {result.get('excel_path')}")
    print(f"Dashboard path : {result.get('dashboard_path')}")

    # Validate JSON content
    if result.get("json_path"):
        with open(result["json_path"]) as f:
            data = json.load(f)
        assert "timestamp" in data, "FAIL: Missing timestamp"
        assert "market_regime" in data, "FAIL: Missing market_regime"
        assert "top_10_stocks" in data, "FAIL: Missing top_10_stocks"
        assert "portfolio" in data, "FAIL: Missing portfolio"
        assert "trade_signals" in data, "FAIL: Missing trade_signals"
        assert "risk_summary" in data, "FAIL: Missing risk_summary"
        print("\n✅ JSON spec validation passed")
        print(f"   market_regime: {data['market_regime']}")
        print(f"   stocks: {[s['ticker'] for s in data['top_10_stocks']]}")

    if result.get("excel_path"):
        print(f"✅ Excel file created ({result['excel_path'].stat().st_size/1024:.1f} KB)")

    if result.get("dashboard_path"):
        print(f"✅ Dashboard created ({result['dashboard_path'].stat().st_size/1024:.1f} KB)")

    print("\n✅ Stage 8 test complete")
